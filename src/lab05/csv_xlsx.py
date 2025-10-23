from __future__ import annotations
import csv
from pathlib import Path

from openpyxl import Workbook
from openpyxl.utils import get_column_letter

__all__ = ["csv_to_xlsx"]

def _ensure_ext(path: Path, *allowed: str) -> None:
    if path.suffix.lower() not in allowed:
        raise ValueError(f"Неверный тип файла для {path.name}: ожидается {', '.join(allowed)}")

def csv_to_xlsx(csv_path: str, xlsx_path: str) -> None:
    """
    Конвертирует CSV в XLSX.
    - Первая строка CSV — заголовок
    - Лист называется "Sheet1"
    - Колонки — автоширина по длине текста (не менее 8 символов)
    """
    cpath = Path(csv_path)
    xpath = Path(xlsx_path)

    _ensure_ext(cpath, ".csv")
    _ensure_ext(xpath, ".xlsx")

    if not cpath.exists():
        raise FileNotFoundError(str(cpath))

    with cpath.open(encoding="utf-8") as f:
        sample = f.read(4096)
        if not sample.strip():
            raise ValueError("Пустой CSV")
        f.seek(0)
        try:
            dialect = csv.Sniffer().sniff(sample)
        except csv.Error:
            dialect = csv.excel

        reader = csv.reader(f, dialect)
        rows = list(reader)
        if not rows:
            raise ValueError("CSV не содержит данных")
        header = rows[0]
        if not header or any(h is None or h == "" for h in header):
            raise ValueError("CSV должен содержать непустой заголовок")

    wb = Workbook()
    ws = wb.active
    ws.title = "Sheet1"

    # Записываем строки и считаем ширины
    max_widths = [max(8, len(str(h))) for h in header]

    for r_idx, row in enumerate(rows, start=1):
        ws.append(row)
        # обновляем ширины начиная со второй строки тоже
        for i, cell in enumerate(row):
            l = len(str(cell)) if cell is not None else 0
            if i >= len(max_widths):
                max_widths.append(max(8, l))
            else:
                max_widths[i] = max(max_widths[i], l)

    # Применяем ширины (+2 символа отступа)
    for i, w in enumerate(max_widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = max(8, w + 2)

    wb.save(xpath)
